# coding: utf-8
#!/usr/bin/python


import importlib
# importlib.reload(sys)

import sys
import json
import xlrd
import openpyxl
import pandas as pd 

__author__ = 'pepper'
__site__ = 'pepper.com'

# 指定信息在 xls 表格内的列数
_colOfClassName = 0
_colOfStartWeek = 1
_colOfEndWeek = 2
_colOfWeekday = 3
_colOfClassTime = 4
_colOfClassroom = 5

def main():
	# 读取 excel 文件
	# 指定engine为openpyxl
	table = pd.read_excel('classInfo.xlsx', engine='openpyxl',sheet_name = 'Sheet1')
	
	#data = xlrd.open_workbook('classInfo.xlsx')
	#data = data.sheets()[0]
	#table = data['Sheet1']
	table = pd.DataFrame(table.values)
	# print table.cell(1,0).value
	# 基础信息
	numOfRow = table.shape[0]  #获取行数,即课程数
	numOfCol = table.shape[1]  #获取列数,即信息量
	headStr = '{\n"classInfo":[\n'
	tailStr = ']\n}'
	classInfoStr = ''
	classInfoArray = []
	# 信息列表
	# lengthOfList = numOfRow-1
	classNameList = []
	startWeekList = []
	endWeekList = []
	weekdayList = []
	classTimeList = []
	classroomList = []

	# 确定配置内容
	info = "\n欢迎使用课程表生成工具·Excel 解析器。\n这是你 Excel 列信息配置，请检查。\n\n如若有误，请自行编辑 excelReader 文件第 12～17 行\n\n"
	info += "ClassName: " + str(_colOfClassName) + "列\n"
	info += "StartWeek: " + str(_colOfStartWeek) + "列\n"
	info += "EndWeek: " + str(_colOfEndWeek) + "列\n"
	info += "Weekday: " + str(_colOfWeekday) + "列\n"
	info += "ClassTime: " + str(_colOfClassTime) + "列\n"
	info += "Classroom: " + str(_colOfClassroom) + "列\n"
	print (info)
	# info += "输入 0 继续，输入 1 退出："
	option = input("输入 0 继续，输入其他内容退出：")
	if option == "1":
		sys.exit()
	

	# 开始操作
	# 将信息加载到列表
	i = 1
	while i < numOfRow :
		index = i-1

		# classNameList.append(((table.cell(i, _colOfClassName).value)))
		# startWeekList.append(str(int((table.cell(i, _colOfStartWeek).value))))
		# endWeekList.append(str(int((table.cell(i, _colOfEndWeek).value))))
		# weekdayList.append(str(int((table.cell(i, _colOfWeekday).value))))
		# classTimeList.append(str(int((table.cell(i, _colOfClassTime).value))))
		# classroomList.append(str(((table.cell(i, _colOfClassroom).value))))

		classNameList.append(table.at[i, _colOfClassName])
		startWeekList.append(str(int(table.at[i, _colOfStartWeek])))
		endWeekList.append(str(int(table.at[i, _colOfEndWeek])))
		weekdayList.append(str(int(table.at[i, _colOfWeekday])))
		classTimeList.append(str(int(table.at[i, _colOfClassTime])))
		classroomList.append(str((table.at[i, _colOfClassroom])))
		
		i += 1
	i = 0
	itemHeadStr = '{\n'
	itemTailStr = '\n}'

	classInfoStr += headStr
	for className in classNameList:
		itemClassInfoStr = ""
		itemClassInfoStr  = itemHeadStr + '"className":"' + className + '",'
		itemClassInfoStr += '"week":{\n"startWeek":' + startWeekList[i] + ',\n'
		itemClassInfoStr += '"endWeek":' + endWeekList[i] + '\n},\n'
		itemClassInfoStr += '"weekday":' + weekdayList[i] + ',\n'
		itemClassInfoStr += '"classTime":' + classTimeList[i] + ',\n'
		itemClassInfoStr += '"classroom":"' + classroomList[i] + '"\n'
		itemClassInfoStr += itemTailStr
		classInfoStr += itemClassInfoStr
		if i!=len(classNameList)-1 :
			classInfoStr += ","
		i += 1
	classInfoStr += tailStr
	# print classInfoStr
	with open('conf_classInfo.json','w') as f:

		f.write(classInfoStr)
		f.close()
	print("\nALL DONE !")

importlib.reload(sys);
# reload(sys);

# sys.setdefaultencoding('utf-8');
main()